# -*- coding: utf-8 -*-
"""
Tencent is pleased to support the open source community by making BK-BASE 蓝鲸基础平台 available.

Copyright (C) 2021 THL A29 Limited, a Tencent company.  All rights reserved.

BK-BASE 蓝鲸基础平台 is licensed under the MIT License.

License for BK-BASE 蓝鲸基础平台:
--------------------------------------------------------------------
Permission is hereby granted, free of charge, to any person obtaining a copy of this software and associated
documentation files (the "Software"), to deal in the Software without restriction, including without limitation
the rights to use, copy, modify, merge, publish, distribute, sublicense, and/or sell copies of the Software,
and to permit persons to whom the Software is furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in all copies or substantial
portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR IMPLIED, INCLUDING BUT NOT
LIMITED TO THE WARRANTIES OF MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN
NO EVENT SHALL THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY,
WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM, OUT OF OR IN CONNECTION WITH THE
SOFTWARE OR THE USE OR OTHER DEALINGS IN THE SOFTWARE.
"""
import io
import json

import xlrd
from common.exceptions import ValidationError
from common.log import logger
from conf import dataapi_settings
from django.utils.translation import ugettext as _
from django.utils.translation import ugettext_noop
from openpyxl import load_workbook

from ...collectors.utils.language import Bilingual
from ...exceptions import CollectorError, CollerctorCode
from ...models import (
    AccessOperationLog,
    AccessRawData,
    AccessRawDataTask,
    AccessResourceInfo,
    AccessTask,
    DatabusChannelClusterConfig,
)
from ..base_collector import BaseAccess, BaseAccessTask
from ..utils import kafka_util
from .serializers import FileResourceSerializer


class FileAccess(BaseAccess):
    """
    文件上传接入
    """

    is_disposable = True
    data_scenario = "file"
    access_serializer = FileResourceSerializer
    task_disable = True

    def create_after(self):
        # 创建执行记录
        task_id = BaseAccessTask.create(
            self.access_param["bk_biz_id"],
            self.access_param["bk_username"],
            self.data_scenario,
            {},
            AccessTask.ACTION.DEPLOY,
        )
        AccessRawDataTask.objects.create(
            task_id=task_id,
            raw_data_id=self.access_param["raw_data_id"],
            created_by=self.access_param["bk_username"],
            updated_by=self.access_param["bk_username"],
            description="",
        )

        task = BaseAccessTask(task_id=task_id)
        resource_list = AccessResourceInfo.objects.filter(raw_data_id=self.access_param["raw_data_id"])
        if not resource_list:
            task.log(
                msg=Bilingual(ugettext_noop(u"不存在部署资源源配置信息")),
                level="ERROR",
                task_log=True,
            )
            task.task.set_status(AccessTask.STATUS.FAILURE)
            raise CollectorError(error_code=CollerctorCode.COLLECTOR_NOT_EXSIT_RESOURCE)
        msg_array = list()
        try:
            task.log(msg=Bilingual(ugettext_noop(u"开始解析文件")), level="INFO", task_log=True)
            for resource in resource_list:
                scope = json.loads(resource.resource)
                logger.info(
                    u"Report to kafka, data_msg: %s,raw_data_id:%s"
                    % (scope.get("file_name"), self.access_param["raw_data_id"])
                )
                raw_data = AccessRawData.objects.get(id=self.access_param["raw_data_id"])
                data_encoding = raw_data.data_encoding
                if scope.get("file_name"):
                    file_name = scope.get("file_name")
                    file_path = "{}/{}".format(
                        dataapi_settings.UPLOAD_MEDIA_ROOT,
                        file_name,
                    )
                    if file_name.endswith(".csv"):
                        msg_array = msg_array + self.csv_paser(file_path, data_encoding)
                    elif file_name.endswith(".xlsx"):
                        msg_array = msg_array + self.xlsx_paser(file_path)
                    elif file_name.endswith(".xls"):
                        msg_array = msg_array + self.xls_paser(file_path)
                    else:
                        task.log(
                            msg=Bilingual(ugettext_noop(u"暂时不支持该文件类型")),
                            level="ERROR",
                            task_log=True,
                        )
                        task.task.set_status(AccessTask.STATUS.FAILURE)
                        raise ValidationError(message=_(u"目前只支持.csv,.xls,.xlsx 文件"))

            # 发送kafka
            task.log(msg=Bilingual(ugettext_noop(u"开始上报数据")), level="INFO", task_log=True)
            storage_channel_id = raw_data.storage_channel_id

            storage = DatabusChannelClusterConfig.objects.get(id=storage_channel_id)

            kafka_bs = storage.cluster_domain + ":" + str(storage.cluster_port)
            topic = raw_data.topic_name
            kafka_util.send_message_to_kafka(kafka_bs, topic, msg_array)
            task.log(msg=Bilingual(ugettext_noop(u"上报数据完成")), level="INFO", task_log=True)
            task.task.set_status(AccessTask.STATUS.SUCCESS)
        except CollectorError as e:
            task.log(
                msg=Bilingual(ugettext_noop(u"The expected file upload abnormal access:{}")).format(str(e)),
                level="ERROR",
                task_log=True,
            )
            self.update_status(AccessOperationLog.STATUS.FAILURE)
            task.task.set_status(AccessTask.STATUS.FAILURE)
            raise e

    def xlsx_paser(self, file_path):
        msg_array = list()
        try:
            wb = load_workbook(filename=file_path)

            sheets = wb.get_sheet_names()
            for sheet in sheets:
                ws = wb.get_sheet_by_name(sheet)
                rows = list(ws.rows)
                for row in rows[1:]:
                    msg_array.append({ws.cell(row=1, column=col.col_idx).value: col.value for col in row})
        except Exception as e:
            logger.error(
                u"xlsx  file access failed,raw_data_id:{}, failed reason: {}, ".format(
                    self.access_param["raw_data_id"], e
                )
            )
            raise CollectorError(
                error_code=CollerctorCode.COLLECTOR_FILE_ACCESS_FAIL,
                message=u"xls文件解析失败,原因:%s" % e,
            )

        return msg_array

    def xls_paser(self, file_path):
        msg_array = list()
        try:
            wb = xlrd.open_workbook(file_path)

            sheets = wb.sheet_names()
            for sheet in sheets:
                ws = wb.sheet_by_name(sheet)
                colnames = ws.row_values(0)
                for rownum in range(1, ws.nrows):
                    row = ws.row_values(rownum)
                    msg_array.append({colnames[i]: row[i] for i in range(len(colnames))})
        except Exception as e:
            logger.error(
                u"xls file access failed,raw_data_id:{}, failed reason: {}, ".format(
                    self.access_param["raw_data_id"], e
                )
            )
            raise CollectorError(
                error_code=CollerctorCode.COLLECTOR_FILE_ACCESS_FAIL,
                message=u"xls文件解析失败:%s" % e,
            )

        return msg_array

    def csv_paser(self, file_path, data_encoding):
        msg_array = list()

        try:
            with io.open(file_path, "rU", encoding=data_encoding) as csvfile:
                title = None
                for line in csvfile.readlines():
                    line = line.strip("\n").strip("\r")
                    if title is None:
                        title = line.split(",")
                    else:
                        row = line.split(",")
                        msg_array.extend([{title[i]: row[i] for i in range(len(row))}])
                msg_array = json.loads(json.dumps(msg_array))
        except Exception as e:
            logger.error(
                u"csv file access failed,raw_data_id:{}, failed reason: {}".format(self.access_param["raw_data_id"], e)
            )
            raise CollectorError(
                error_code=CollerctorCode.COLLECTOR_FILE_ACCESS_FAIL,
                message=u"csv文件解析失败,原因:%s" % e,
            )

        return msg_array
